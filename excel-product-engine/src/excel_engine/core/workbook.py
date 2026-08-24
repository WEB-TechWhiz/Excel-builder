"""Product-agnostic workbook abstraction wrapping openpyxl.Workbook.

This is the single entry point every product/component uses to build a
workbook. It knows how to create/find/remove sheets and save/load the
file — it never contains business logic about any specific product
(that lives in ``products/<name>``).

    >>> workbook = ExcelWorkbook()
    >>> dashboard = workbook.add_sheet("Dashboard")
    >>> dashboard.set_value("A1", "Financial OS")
    >>> workbook.save("output/financial_os.xlsx")
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook as _OpenpyxlWorkbook
from openpyxl import load_workbook as _openpyxl_load_workbook

from excel_engine.core.metadata import WorkbookMetadata
from excel_engine.core.worksheet import Worksheet
from excel_engine.exceptions.errors import ExportError, SheetNotFoundError
from excel_engine.logging_config import get_logger

logger = get_logger("core.workbook")


class ExcelWorkbook:
    """A typed, product-agnostic Excel workbook under construction."""

    def __init__(self, metadata: WorkbookMetadata | None = None) -> None:
        self._raw = _OpenpyxlWorkbook()
        self._sheets: dict[str, Worksheet] = {}
        self.metadata = metadata

        # openpyxl always creates one default sheet ("Sheet"). We track its
        # title and quietly repurpose it the first time add_sheet() is
        # called, so callers of this class never see a stray extra tab.
        self._default_sheet_pending = True

        if metadata:
            metadata.apply(self._raw)

    # -- sheet lifecycle ----------------------------------------------------
    def add_sheet(self, name: str, index: int | None = None) -> Worksheet:
        if self._default_sheet_pending:
            raw_sheet = self._raw.active
            raw_sheet.title = name
            self._default_sheet_pending = False
        else:
            raw_sheet = self._raw.create_sheet(title=name, index=index)

        sheet = Worksheet(raw_sheet)
        self._sheets[name] = sheet
        logger.debug("Added sheet %r", name)
        return sheet

    def get_sheet(self, name: str) -> Worksheet:
        try:
            return self._sheets[name]
        except KeyError as exc:
            raise SheetNotFoundError(f"No sheet named {name!r} in this workbook") from exc

    def remove_sheet(self, name: str) -> None:
        sheet = self.get_sheet(name)
        self._raw.remove(sheet.raw)
        del self._sheets[name]
        logger.debug("Removed sheet %r", name)

    def has_sheet(self, name: str) -> bool:
        return name in self._sheets

    @property
    def sheet_names(self) -> list[str]:
        """Sheet names in their actual on-disk/tab order.

        Reads from the underlying openpyxl workbook (the real source of
        truth for order) rather than the internal lookup dict, so this
        stays correct after `reorder_sheet()` — see
        `test_sheet_names_reflects_reorder`.
        """
        return list(self._raw.sheetnames)

    def reorder_sheet(self, name: str, index: int) -> None:
        sheet = self.get_sheet(name)
        current_index = self._raw.index(sheet.raw)
        self._raw.move_sheet(sheet.raw, offset=index - current_index)

    # -- persistence ----------------------------------------------------
    def save(self, path: str | Path) -> Path:
        if not self._sheets:
            raise ExportError("Cannot save a workbook with no sheets.")
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            self._raw.save(str(output_path))
        except OSError as exc:
            raise ExportError(f"Failed to save workbook to {output_path}: {exc}") from exc
        logger.info("Workbook saved to %s", output_path)
        return output_path

    @classmethod
    def load(cls, path: str | Path) -> ExcelWorkbook:
        """Load an existing .xlsx back into the typed wrapper.

        Note: openpyxl reads formulas as text by default, not their
        calculated values (see section 32 — it does not evaluate Excel
        formulas itself).
        """
        raw = _openpyxl_load_workbook(str(path))
        instance = cls.__new__(cls)
        instance._raw = raw
        instance._sheets = {name: Worksheet(raw[name]) for name in raw.sheetnames}
        instance._default_sheet_pending = False
        instance.metadata = None
        return instance

    @property
    def raw(self) -> _OpenpyxlWorkbook:
        """Escape hatch to the underlying openpyxl.Workbook."""
        return self._raw

    def __repr__(self) -> str:
        return f"ExcelWorkbook(sheets={self.sheet_names!r})"
