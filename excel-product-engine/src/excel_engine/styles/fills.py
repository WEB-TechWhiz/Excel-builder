"""Fill presets — the fill half of the design system."""

from __future__ import annotations

from openpyxl.styles import PatternFill

NO_FILL = PatternFill(fill_type=None)


def solid_fill(hex_color: str) -> PatternFill:
    """A solid background fill from a 6-digit hex color (no '#')."""
    return PatternFill(fill_type="solid", fgColor=hex_color)
