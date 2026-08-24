"""StyleManager — the single place product/component code asks for
ready-to-use openpyxl style objects (Font, PatternFill, Border,
Alignment) for a given theme, instead of constructing them by hand.

    >>> style = StyleManager.for_theme("premium")
    >>> cell.font = style.title_font
    >>> cell.fill = style.header_fill
"""

from __future__ import annotations

from dataclasses import dataclass
from enum import IntEnum

from openpyxl.styles import Alignment, Border, Font, PatternFill

from excel_engine.config.themes import AVAILABLE_THEMES, is_valid_theme
from excel_engine.exceptions.errors import ProductConfigurationError
from excel_engine.styles import alignment as _alignment
from excel_engine.styles import borders as _borders
from excel_engine.styles import fills as _fills
from excel_engine.styles.colors import ColorPalette, get_palette
from excel_engine.styles.fonts import DEFAULT_TYPOGRAPHY, FontSpec, Typography


class Spacing(IntEnum):
    """Abstract layout scale for later phases (components) to consume —
    e.g. as row/column pitch when positioning KPI cards. Not an Excel
    unit itself; each component decides how to interpret it (rows,
    columns, or points) for its own layout.
    """

    XS = 4
    SM = 8
    MD = 12
    LG = 16
    XL = 24
    XXL = 32


@dataclass(frozen=True, slots=True)
class StyleManager:
    """A resolved theme: a color palette + typography, exposed as ready
    openpyxl style objects for every semantic purpose components need.
    """

    theme_name: str
    palette: ColorPalette
    typography: Typography = DEFAULT_TYPOGRAPHY

    @classmethod
    def for_theme(cls, theme_name: str, typography: Typography | None = None) -> StyleManager:
        if not is_valid_theme(theme_name):
            raise ProductConfigurationError(
                f"Unknown theme {theme_name!r}. Available: {AVAILABLE_THEMES}"
            )
        return cls(
            theme_name=theme_name,
            palette=get_palette(theme_name),
            typography=typography or DEFAULT_TYPOGRAPHY,
        )

    # -- fonts ------------------------------------------------------------
    def _font(self, spec: FontSpec, color: str, bold: bool | None = None) -> Font:
        return Font(
            name=self.typography.font_family,
            size=spec.size,
            bold=spec.bold if bold is None else bold,
            italic=spec.italic,
            color=color,
        )

    @property
    def title_font(self) -> Font:
        """For the big banner title on a Dashboard sheet — on-primary color."""
        return self._font(self.typography.title, self.palette.on_primary)

    @property
    def heading_font(self) -> Font:
        return self._font(self.typography.heading, self.palette.text)

    @property
    def subheading_font(self) -> Font:
        return self._font(self.typography.subheading, self.palette.muted)

    @property
    def body_font(self) -> Font:
        return self._font(self.typography.body, self.palette.text)

    @property
    def caption_font(self) -> Font:
        return self._font(self.typography.caption, self.palette.muted)

    @property
    def header_font(self) -> Font:
        """Table header row text — heading weight, on-primary color."""
        return self._font(self.typography.heading, self.palette.on_primary)

    @property
    def kpi_label_font(self) -> Font:
        return self._font(self.typography.caption, self.palette.primary, bold=True)

    @property
    def kpi_value_font(self) -> Font:
        return self._font(self.typography.heading, self.palette.primary, bold=True)

    @property
    def input_font(self) -> Font:
        """User-editable data — always blue, independent of theme, matching
        the universal spreadsheet convention (blue = input, black = formula).
        """
        return Font(name=self.typography.font_family, size=self.typography.body.size, color="0000FF")

    @property
    def formula_font(self) -> Font:
        """Auto-calculated data — theme text color, never the input blue."""
        return self._font(self.typography.body, self.palette.text)

    @property
    def subtitle_font(self) -> Font:
        """Small caption text painted over a primary-colored fill (e.g. a
        title banner's subtitle line) — on-primary color, not muted gray.
        """
        return self._font(self.typography.caption, self.palette.on_primary)

    @property
    def nav_link_font(self) -> Font:
        """A clickable navbar item — underlined, primary color."""
        return Font(
            name=self.typography.font_family,
            size=self.typography.body.size,
            color=self.palette.primary,
            underline="single",
        )

    @property
    def nav_active_font(self) -> Font:
        """The current sheet in a navbar — bold, no underline (not a link)."""
        return Font(
            name=self.typography.font_family,
            size=self.typography.body.size,
            color=self.palette.text,
            bold=True,
        )

    # -- fills ------------------------------------------------------------
    @property
    def header_fill(self) -> PatternFill:
        return _fills.solid_fill(self.palette.primary)

    @property
    def card_fill(self) -> PatternFill:
        return _fills.solid_fill(self.palette.surface)

    @property
    def no_fill(self) -> PatternFill:
        return _fills.NO_FILL

    # -- borders ------------------------------------------------------------
    @property
    def thin_border(self) -> Border:
        return _borders.box(color=self.palette.muted, style=_borders.THIN)

    # -- alignment ------------------------------------------------------------
    @property
    def center(self) -> Alignment:
        return _alignment.CENTER

    @property
    def left(self) -> Alignment:
        return _alignment.LEFT

    @property
    def right(self) -> Alignment:
        return _alignment.RIGHT

    @property
    def wrap(self) -> Alignment:
        return _alignment.WRAP
