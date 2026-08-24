"""Typed data tables — a column schema (text/number/currency/percent/
date/list) layered on top of the purely visual
`excel_engine.components.add_data_table`, wiring in number formats and
dropdown validation for "list" columns, plus optional same-row formula
columns (e.g. "Gain/Loss" = Current Value - Quantity*Purchase Price).

This is deliberately a separate module from `components.tables`: that
one only knows how to draw a banded table, this one knows what a
column's *type* means (see docs/architecture.md's Phase 4 notes).
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Literal

from openpyxl.utils import get_column_letter

from excel_engine.components.tables import add_data_table
from excel_engine.core.cell import CellAddress
from excel_engine.core.range import CellRange
from excel_engine.core.worksheet import Worksheet
from excel_engine.data.validation import add_dropdown
from excel_engine.exceptions.errors import WorkbookBuildError
from excel_engine.styles.style_manager import StyleManager

ColumnType = Literal["text", "number", "currency", "percent", "date", "list"]

_NUMBER_FORMAT_TEMPLATES = {
    "currency": '"{sym}"#,##0',
    "number": "#,##0",
    "percent": "0.0%",
    "date": "dd-mmm-yyyy",
    "text": "General",
    "list": "General",
}

# Minimum width (characters) needed for each type's formatted value to
# display without Excel's "###" truncation indicator — overrides
# components.add_data_table's generic header-length fallback now that
# real column types are known.
_MIN_COLUMN_WIDTH = {
    "currency": 14,
    "number": 12,
    "percent": 10,
    "date": 13,
    "text": 18,
    "list": 16,
}

_PLACEHOLDER_RE = re.compile(r"\{([^}]+)\}")


@dataclass(frozen=True, slots=True)
class ColumnSchema:
    header: str
    type: ColumnType = "text"
    options: tuple[str, ...] = ()
    formula: str | None = None
    """Template using {Other Header} placeholders, e.g.
    '{Current Value}-{Quantity}*{Purchase Price}'. Resolved per-row into
    a real formula, blank until the first referenced column has a
    value. Column order doesn't matter — placeholders can reference any
    other column in the same table, including ones defined later.
    """

    def __post_init__(self) -> None:
        if self.type == "list" and not self.options:
            raise WorkbookBuildError(f"Column {self.header!r} is type='list' but has no options")
        if self.formula and self.type == "list":
            raise WorkbookBuildError(f"Column {self.header!r} can't be both type='list' and formula-driven")


@dataclass(frozen=True, slots=True)
class TypedTable:
    """Everything downstream code (the formula engine, KPI cards) needs
    to reference this table's columns correctly — the typed evolution
    of the ad-hoc `_TableMeta` pattern used before this module existed.
    """

    sheet_name: str
    table_name: str
    column_letters: dict[str, str]
    first_data_row: int
    last_data_row: int


def _resolve_row_formula(template: str, column_letters: dict[str, str], row: int) -> str:
    """Turn '{Quantity}*{Purchase Price}' into a formula for this row,
    blank until the first referenced column is filled in:
    '=IF(D5="","",D5*E5)'.
    """
    refs_used: list[str] = []

    def repl(match: re.Match[str]) -> str:
        letter = column_letters[match.group(1)]
        refs_used.append(letter)
        return f"{letter}{row}"

    expr = _PLACEHOLDER_RE.sub(repl, template)
    trigger = refs_used[0] if refs_used else None
    return f'=IF({trigger}{row}="","",{expr})' if trigger else f"={expr}"


def add_typed_table(
    ws: Worksheet,
    style: StyleManager,
    columns: list[ColumnSchema],
    n_rows: int,
    table_name: str,
    top_left: str | CellAddress = "A1",
    currency_symbol: str = "₹",
) -> TypedTable:
    anchor = top_left if isinstance(top_left, CellAddress) else CellAddress.from_a1(top_left)
    headers = [c.header for c in columns]

    full_range = add_data_table(ws, style, headers, n_rows, table_name, top_left=anchor)

    number_formats = {
        key: (tmpl.format(sym=currency_symbol) if "{sym}" in tmpl else tmpl)
        for key, tmpl in _NUMBER_FORMAT_TEMPLATES.items()
    }

    first_data_row = anchor.row + 1
    last_data_row = full_range.end.row
    column_letters: dict[str, str] = {
        col.header: get_column_letter(anchor.column + i) for i, col in enumerate(columns)
    }

    for i, col in enumerate(columns):
        col_index = anchor.column + i
        letter = column_letters[col.header]

        min_width = _MIN_COLUMN_WIDTH.get(col.type, 12)
        header_width = len(col.header) + 4
        ws.set_column_width(letter, max(min_width, header_width))

        for row in range(first_data_row, last_data_row + 1):
            cell = ws.raw.cell(row=row, column=col_index)
            cell.number_format = number_formats[col.type]
            if col.formula:
                cell.value = _resolve_row_formula(col.formula, column_letters, row)
                cell.font = style.formula_font

        if col.type == "list":
            data_range = CellRange(
                start=CellAddress(row=first_data_row, column=col_index),
                end=CellAddress(row=last_data_row, column=col_index),
            )
            add_dropdown(ws, data_range, list(col.options))

    return TypedTable(
        sheet_name=ws.name,
        table_name=table_name,
        column_letters=column_letters,
        first_data_row=first_data_row,
        last_data_row=last_data_row,
    )
