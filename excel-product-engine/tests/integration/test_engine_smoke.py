"""End-to-end smoke test: build a small workbook purely through the typed
engine API (no direct openpyxl calls from the test) and confirm the
result is a valid, correct .xlsx on disk — not just something our own
wrapper thinks it can read back.
"""

from openpyxl import load_workbook

from excel_engine.core.metadata import WorkbookMetadata
from excel_engine.core.workbook import ExcelWorkbook


def test_minimal_workbook_end_to_end(tmp_path):
    metadata = WorkbookMetadata(title="Financial OS", author="MuffinCodes", version="0.1.0")
    workbook = ExcelWorkbook(metadata=metadata)

    dashboard = workbook.add_sheet("Dashboard")
    dashboard.set_value("A1", "Financial OS")
    dashboard.merge("A1:C1")  # title banner — B1/C1 become MergedCells, by design
    dashboard.set_formula("A2", "=1+1")  # outside the merge, so its value survives
    dashboard.freeze_panes("A3")
    dashboard.set_tab_color("1F4E78")

    income = workbook.add_sheet("Income")
    income.set_value("A1", "Source")

    out_path = workbook.save(tmp_path / "financial_os.xlsx")

    raw = load_workbook(str(out_path))
    assert raw.sheetnames == ["Dashboard", "Income"]
    assert raw["Dashboard"]["A1"].value == "Financial OS"
    assert raw["Dashboard"]["A2"].value == "=1+1"
    assert raw.properties.title == "Financial OS"
    assert raw.properties.creator == "MuffinCodes"
