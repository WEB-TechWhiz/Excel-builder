"""Proves Phase 1 (ProductConfig), Phase 2 (ExcelWorkbook) and Phase 3
(StyleManager) actually connect — not just that each passes its own
isolated unit tests.
"""

from openpyxl import load_workbook

from excel_engine.config.product_config import ProductConfig
from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.styles.style_manager import StyleManager


def test_product_config_theme_drives_real_workbook_styling(tmp_path):
    config = ProductConfig(
        name="Financial OS",
        version="1.0.0",
        author="MuffinCodes",
        theme={"name": "premium"},
    )
    style = StyleManager.for_theme(config.theme.name)

    workbook = ExcelWorkbook()
    dashboard = workbook.add_sheet("Dashboard")
    dashboard.set_value("A1", "Financial OS")
    dashboard.raw["A1"].font = style.title_font
    dashboard.raw["A1"].fill = style.header_fill
    dashboard.raw["A1"].alignment = style.center

    out_path = workbook.save(tmp_path / "styled.xlsx")

    # Read back with raw openpyxl — proves the styling is really in the
    # .xlsx bytes, not just held in Python objects that were never saved.
    raw = load_workbook(str(out_path))
    cell = raw["Dashboard"]["A1"]
    assert cell.font.bold is True
    assert cell.font.color.rgb.endswith(style.palette.on_primary)
    assert cell.fill.fgColor.rgb.endswith(style.palette.primary)
    assert cell.alignment.horizontal == "center"
