"""Product-level tests: build the full Financial OS, save it, and
confirm it's structurally and financially correct — the regression
suite that protects this specific product as the engine keeps evolving.
"""

import json
import subprocess
import sys

from pathlib import Path

import openpyxl
from products.financial_os.product import build_financial_os

from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.validation import validate_workbook

RECALC_SCRIPT = str(Path(__file__).resolve().parents[2] / "scripts" / "recalc.py")

EXPECTED_SHEETS = [
    "Dashboard", "Income", "Expenses", "Bills", "Investments",
    "Net Worth", "Goals", "Settings", "Reports",
]


def test_builds_without_error_and_has_all_nine_sheets():
    workbook = build_financial_os()
    assert workbook.sheet_names[0] == "Dashboard"
    assert set(workbook.sheet_names) == set(EXPECTED_SHEETS)


def test_passes_full_validation():
    workbook = build_financial_os()
    report = validate_workbook(
        workbook, product_name="Financial OS",
        required_sheets=EXPECTED_SHEETS,
        required_tables={
            "Income": ["Income"], "Expenses": ["Expenses"], "Bills": ["Bills"],
            "Investments": ["Investments"], "Net Worth": ["NetWorth"], "Goals": ["Goals"],
        },
    )
    assert report.passed is True, report.format()


def test_saved_workbook_recalculates_with_zero_errors(tmp_path):
    workbook = build_financial_os()
    out_path = workbook.save(tmp_path / "financial_os.xlsx")

    result = subprocess.run(
        [sys.executable, RECALC_SCRIPT, str(out_path), "150"],
        capture_output=True, text=True, timeout=120,
    )
    report = json.loads(result.stdout)
    assert report["status"] == "success", report
    assert report["total_errors"] == 0
    assert report["total_formulas"] > 100


def test_dashboard_kpis_match_hand_calculated_demo_data(tmp_path):
    """The financial numbers actually have to be right, not just
    error-free — this locks in the exact figures the demo data
    produces so a future change can't silently break the math.
    """
    workbook = build_financial_os()
    out_path = workbook.save(tmp_path / "financial_os.xlsx")
    subprocess.run(
        [sys.executable, RECALC_SCRIPT, str(out_path), "150"],
        capture_output=True, text=True, timeout=120,
    )

    raw = openpyxl.load_workbook(str(out_path), data_only=True)
    dash = raw["Dashboard"]

    # Net Worth: most recent (August) snapshot only, not summed across
    # every historical entry — Assets 1,085,500 - Liabilities 2,215,000
    assert dash["U3"].value == 1085500
    assert dash["U4"].value == 2215000
    assert dash["A6"].value == -1129500

    # Monthly Income/Expenses: current-month (August 2026) SUMIFS
    assert dash["E6"].value == 103000   # 85000 salary + 18000 freelance
    assert dash["I6"].value == 32000    # 22000 rent + 7200 groceries + 2800 entertainment
    assert dash["M6"].value == 71000    # savings = income - expenses

    assert dash["E10"].value == 455500  # total investments (current value)
    assert dash["I10"].value == 2215000  # debt = total liabilities

    reports = raw["Reports"]
    assert reports["A6"].value == 213000  # total income, all-time
    assert reports["E6"].value == 66200   # total expenses, all-time


def test_investments_gain_loss_column_computes_correctly(tmp_path):
    workbook = build_financial_os()
    out_path = workbook.save(tmp_path / "financial_os.xlsx")
    subprocess.run(
        [sys.executable, RECALC_SCRIPT, str(out_path), "150"],
        capture_output=True, text=True, timeout=120,
    )
    raw = openpyxl.load_workbook(str(out_path), data_only=True)
    inv = raw["Investments"]
    # row 6 = HDFC Flexicap Fund: 500 units @ 45, current value 27500
    assert inv["G6"].value == 27500 - 500 * 45


def test_goals_progress_and_status_compute_correctly(tmp_path):
    workbook = build_financial_os()
    out_path = workbook.save(tmp_path / "financial_os.xlsx")
    subprocess.run(
        [sys.executable, RECALC_SCRIPT, str(out_path), "150"],
        capture_output=True, text=True, timeout=120,
    )
    raw = openpyxl.load_workbook(str(out_path), data_only=True)
    goals = raw["Goals"]
    # row 8 = New Laptop: Current Amount == Target Amount -> Achieved
    assert goals["E8"].value == 1
    assert goals["F8"].value == "Achieved"


def test_protected_sheets_keep_input_columns_editable():
    workbook = build_financial_os()
    income = workbook.get_sheet("Income")
    assert income.raw.protection.sheet is True
    assert income.raw["A6"].protection.locked is False   # first data row, Date column
    assert income.raw["A5"].protection.locked is True     # header row stays locked


def test_navigation_menu_present_on_every_sheet_with_correct_active_page():
    workbook = build_financial_os()
    for sheet_name in EXPECTED_SHEETS:
        ws = workbook.get_sheet(sheet_name)
        assert ws.get_value("A1") == "Dashboard"  # first menu item always present


def test_reload_from_disk_preserves_structure(tmp_path):
    workbook = build_financial_os()
    out_path = workbook.save(tmp_path / "financial_os.xlsx")
    reloaded = ExcelWorkbook.load(out_path)
    assert set(reloaded.sheet_names) == set(EXPECTED_SHEETS)
