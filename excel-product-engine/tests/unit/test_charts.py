from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart

from excel_engine.charts import add_bar_chart, add_doughnut_chart, add_line_chart, add_pie_chart
from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.data.tables import ColumnSchema, add_typed_table
from excel_engine.styles.style_manager import StyleManager


def _orders_sheet(wb: ExcelWorkbook, style: StyleManager):
    ws = wb.add_sheet("Orders")
    columns = [
        ColumnSchema(header="Date", type="date"),
        ColumnSchema(header="Colorway", type="list", options=("Amber", "Teal")),
        ColumnSchema(header="Amount", type="currency"),
    ]
    add_typed_table(ws, style, columns, n_rows=20, table_name="Orders")
    return ws


def test_bar_chart_embeds_on_sheet():
    wb = ExcelWorkbook()
    style = StyleManager.for_theme("premium")
    ws = _orders_sheet(wb, style)
    add_bar_chart(
        ws, style, "Revenue by Colorway", "Orders", "Colorway", "Amount",
        ["Amber", "Teal"], anchor_cell="B10", data_top_left="T5",
    )
    assert len(ws.raw._charts) == 1
    assert isinstance(ws.raw._charts[0], BarChart)
    assert ws.raw._charts[0].title is not None


def test_line_chart_embeds_and_disables_smoothing():
    wb = ExcelWorkbook()
    style = StyleManager.for_theme("premium")
    ws = _orders_sheet(wb, style)
    add_line_chart(
        ws, style, "Revenue Trend", "Orders", "Date", "Amount",
        anchor_cell="B10", data_top_left="T5", periods=3,
    )
    assert len(ws.raw._charts) == 1
    chart = ws.raw._charts[0]
    assert isinstance(chart, LineChart)
    assert all(s.smooth is False for s in chart.series)


def test_pie_chart_embeds_on_sheet():
    wb = ExcelWorkbook()
    style = StyleManager.for_theme("premium")
    ws = _orders_sheet(wb, style)
    add_pie_chart(
        ws, style, "Share by Colorway", "Orders", "Colorway", "Amount",
        ["Amber", "Teal"], anchor_cell="B10", data_top_left="T5",
    )
    assert len(ws.raw._charts) == 1
    assert isinstance(ws.raw._charts[0], PieChart)


def test_doughnut_chart_embeds_on_sheet():
    wb = ExcelWorkbook()
    style = StyleManager.for_theme("premium")
    ws = _orders_sheet(wb, style)
    add_doughnut_chart(
        ws, style, "Share by Colorway", "Orders", "Colorway", "Amount",
        ["Amber", "Teal"], anchor_cell="B10", data_top_left="T5",
    )
    assert len(ws.raw._charts) == 1
    assert isinstance(ws.raw._charts[0], DoughnutChart)


def test_multiple_charts_on_same_sheet():
    wb = ExcelWorkbook()
    style = StyleManager.for_theme("premium")
    ws = _orders_sheet(wb, style)
    add_bar_chart(ws, style, "Bar", "Orders", "Colorway", "Amount", ["Amber", "Teal"],
                  anchor_cell="B10", data_top_left="T5")
    add_pie_chart(ws, style, "Pie", "Orders", "Colorway", "Amount", ["Amber", "Teal"],
                  anchor_cell="J10", data_top_left="T10")
    assert len(ws.raw._charts) == 2
