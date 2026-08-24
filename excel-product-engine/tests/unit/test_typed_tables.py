import pytest

from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.data.tables import ColumnSchema, add_typed_table
from excel_engine.exceptions.errors import WorkbookBuildError
from excel_engine.styles.style_manager import StyleManager


def _columns():
    return [
        ColumnSchema(header="Date", type="date"),
        ColumnSchema(header="Colorway", type="list", options=("Amber", "Teal")),
        ColumnSchema(header="Amount", type="currency"),
    ]


def test_list_column_without_options_rejected():
    with pytest.raises(WorkbookBuildError):
        ColumnSchema(header="Colorway", type="list")


def test_add_typed_table_returns_correct_metadata():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    style = StyleManager.for_theme("premium")
    table = add_typed_table(ws, style, _columns(), n_rows=20, table_name="Orders")

    assert table.sheet_name == "Orders"
    assert table.table_name == "Orders"
    assert table.column_letters == {"Date": "A", "Colorway": "B", "Amount": "C"}
    assert table.first_data_row == 2
    assert table.last_data_row == 21


def test_number_formats_applied_per_column_type():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    style = StyleManager.for_theme("premium")
    add_typed_table(ws, style, _columns(), n_rows=5, table_name="Orders", currency_symbol="₹")

    assert ws.raw["A2"].number_format == "dd-mmm-yyyy"
    assert ws.raw["C2"].number_format == '"₹"#,##0'


def test_column_widths_are_wide_enough_for_formatted_values():
    """Regression test: a date/currency column must be wide enough that
    Excel doesn't render '###' — found by actually looking at a
    generated workbook, not by inspecting code.
    """
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    style = StyleManager.for_theme("premium")
    add_typed_table(ws, style, _columns(), n_rows=5, table_name="Orders")

    assert ws.raw.column_dimensions["A"].width >= 13  # Date
    assert ws.raw.column_dimensions["C"].width >= 14  # Amount (currency)


def test_list_column_gets_dropdown_validation():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    style = StyleManager.for_theme("premium")
    add_typed_table(ws, style, _columns(), n_rows=5, table_name="Orders")

    dvs = list(ws.raw.data_validations.dataValidation)
    assert len(dvs) == 1
    assert dvs[0].type == "list"
    assert "Amber" in dvs[0].formula1


def test_real_excel_table_is_created():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    style = StyleManager.for_theme("premium")
    add_typed_table(ws, style, _columns(), n_rows=5, table_name="Orders")
    assert "Orders" in ws.raw.tables


def test_formula_column_and_list_type_together_rejected():
    with pytest.raises(WorkbookBuildError):
        ColumnSchema(header="X", type="list", options=("A",), formula="{Y}*2")


def test_formula_column_writes_row_formulas():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Investments")
    style = StyleManager.for_theme("premium")
    columns = [
        ColumnSchema(header="Quantity", type="number"),
        ColumnSchema(header="Purchase Price", type="currency"),
        ColumnSchema(header="Current Value", type="currency"),
        ColumnSchema(header="Gain/Loss", type="currency",
                      formula="{Current Value}-{Quantity}*{Purchase Price}"),
    ]
    table = add_typed_table(ws, style, columns, n_rows=5, table_name="Investments")

    formula = ws.get_value("D2")
    assert formula == '=IF(C2="","",C2-A2*B2)'
    assert table.column_letters == {
        "Quantity": "A", "Purchase Price": "B", "Current Value": "C", "Gain/Loss": "D",
    }


def test_formula_column_uses_formula_font():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Investments")
    style = StyleManager.for_theme("premium")
    columns = [
        ColumnSchema(header="Quantity", type="number"),
        ColumnSchema(header="Gain/Loss", type="currency", formula="{Quantity}*2"),
    ]
    add_typed_table(ws, style, columns, n_rows=3, table_name="Investments")
    assert ws.raw["B2"].font.color.rgb == style.formula_font.color.rgb


def test_formula_column_survives_save_and_reload_and_recalculates(tmp_path):
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Investments")
    style = StyleManager.for_theme("premium")
    columns = [
        ColumnSchema(header="Quantity", type="number"),
        ColumnSchema(header="Purchase Price", type="currency"),
        ColumnSchema(header="Current Value", type="currency"),
        ColumnSchema(header="Gain/Loss", type="currency",
                      formula="{Current Value}-{Quantity}*{Purchase Price}"),
    ]
    add_typed_table(ws, style, columns, n_rows=5, table_name="Investments")
    ws.set_value("A2", 10)
    ws.set_value("B2", 100)
    ws.set_value("C2", 1200)
    out_path = wb.save(tmp_path / "investments.xlsx")

    import json
    import subprocess
    import sys
    from pathlib import Path

    recalc_script = str(Path(__file__).resolve().parents[2] / "scripts" / "recalc.py")
    result = subprocess.run(
        [sys.executable, recalc_script, str(out_path), "5"],
        capture_output=True, text=True, timeout=60,
    )
    report = json.loads(result.stdout)
    assert report["status"] == "success"
    assert report["total_errors"] == 0

    import openpyxl
    reloaded = openpyxl.load_workbook(str(out_path), data_only=True)
    assert reloaded["Investments"]["D2"].value == 200  # 1200 - 10*100
    assert reloaded["Investments"]["D3"].value in (None, "")  # blank row stays blank
