from datetime import date

from openpyxl import load_workbook

from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.data.validation import (
    add_date_range,
    add_dropdown,
    add_number_range,
    add_required,
)


def test_dropdown_registers_list_validation():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    add_dropdown(ws, "B2:B50", ["Amber", "Teal", "Plum"])
    dvs = list(ws.raw.data_validations.dataValidation)
    assert len(dvs) == 1
    assert dvs[0].type == "list"
    assert "Amber" in dvs[0].formula1


def test_number_range_validation():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    add_number_range(ws, "C2:C50", minimum=0, maximum=100000)
    dv = list(ws.raw.data_validations.dataValidation)[0]
    assert dv.type == "decimal"
    assert dv.formula1 == "0"
    assert dv.formula2 == "100000"


def test_date_range_validation():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    add_date_range(ws, "A2:A50", earliest=date(2026, 1, 1), latest=date(2026, 12, 31))
    dv = list(ws.raw.data_validations.dataValidation)[0]
    assert dv.type == "date"
    assert dv.formula1 == "2026-01-01"


def test_required_uses_custom_non_blank_formula():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    add_required(ws, "A2:A50")
    dv = list(ws.raw.data_validations.dataValidation)[0]
    assert dv.type == "custom"
    assert dv.formula1 == 'A2<>""'
    assert dv.allowBlank is False


def test_validations_survive_save_and_reload(tmp_path):
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    add_dropdown(ws, "B2:B10", ["Amber", "Teal"])
    out_path = wb.save(tmp_path / "orders.xlsx")

    raw = load_workbook(str(out_path))
    dvs = list(raw["Orders"].data_validations.dataValidation)
    assert len(dvs) == 1
    assert dvs[0].type == "list"
