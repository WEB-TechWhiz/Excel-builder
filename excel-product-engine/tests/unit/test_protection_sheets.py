from openpyxl import load_workbook

from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.protection import apply_standard_protection
from excel_engine.protection.sheets import protect_sheet, unprotect_sheet


def test_protect_sheet_turns_on_protection():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    protect_sheet(ws)
    assert ws.raw.protection.sheet is True


def test_protect_sheet_does_not_block_formatting_sorting_autofilter():
    """This is the "don't make it unusably restrictive" requirement —
    openpyxl's own defaults (all True = blocked) would fail this if we
    forgot to override them.
    """
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    protect_sheet(ws)
    assert ws.raw.protection.formatCells is False
    assert ws.raw.protection.formatColumns is False
    assert ws.raw.protection.formatRows is False
    assert ws.raw.protection.sort is False
    assert ws.raw.protection.autoFilter is False


def test_unprotect_sheet_turns_it_back_off():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    protect_sheet(ws)
    unprotect_sheet(ws)
    assert ws.raw.protection.sheet is False


def test_protect_sheet_with_password():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    protect_sheet(ws, password="secret")
    assert ws.raw.protection.password is not None  # stored as a hash, not plaintext


def test_apply_standard_protection_unlocks_then_protects():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    apply_standard_protection(ws, editable_ranges=["A2:C51"])

    assert ws.raw.protection.sheet is True
    assert ws.raw["A2"].protection.locked is False    # editable range: unlocked
    assert ws.raw["A1"].protection.locked is True      # header row: still locked


def test_protection_survives_save_and_reload(tmp_path):
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    apply_standard_protection(ws, editable_ranges=["A2:C51"])
    out_path = wb.save(tmp_path / "orders.xlsx")

    raw = load_workbook(str(out_path))
    reloaded_ws = raw["Orders"]
    assert reloaded_ws.protection.sheet is True
    assert reloaded_ws.protection.formatCells is False
    assert reloaded_ws["A2"].protection.locked is False
    assert reloaded_ws["A1"].protection.locked is True
