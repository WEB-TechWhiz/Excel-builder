import pytest
from openpyxl import load_workbook

from excel_engine.components.tables import add_data_table
from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.exceptions.errors import WorkbookBuildError
from excel_engine.styles.style_manager import StyleManager


@pytest.fixture
def sheet():
    wb = ExcelWorkbook()
    return wb.add_sheet("Orders")


def test_table_writes_headers(sheet):
    style = StyleManager.for_theme("premium")
    rng = add_data_table(sheet, style, ["Date", "Amount"], n_rows=10, table_name="Orders")
    assert sheet.get_value("A1") == "Date"
    assert sheet.get_value("B1") == "Amount"
    assert rng.to_a1() == "A1:B11"


def test_table_header_styling(sheet):
    style = StyleManager.for_theme("premium")
    add_data_table(sheet, style, ["Date", "Amount"], n_rows=5, table_name="Orders")
    header_cell = sheet.raw["A1"]
    assert header_cell.fill.fgColor.rgb.endswith(style.palette.primary)
    assert header_cell.font.color.rgb.endswith(style.palette.on_primary)


def test_table_rejects_empty_headers(sheet):
    style = StyleManager.for_theme("premium")
    with pytest.raises(WorkbookBuildError):
        add_data_table(sheet, style, [], n_rows=5, table_name="Orders")


def test_table_rejects_zero_rows(sheet):
    style = StyleManager.for_theme("premium")
    with pytest.raises(WorkbookBuildError):
        add_data_table(sheet, style, ["Date"], n_rows=0, table_name="Orders")


def test_table_name_is_sanitized(sheet):
    style = StyleManager.for_theme("premium")
    add_data_table(sheet, style, ["Date"], n_rows=3, table_name="Order Data!")
    table_names = list(sheet.raw.tables.keys())
    assert table_names == ["Order_Data_"]


def test_table_survives_save_and_reload_as_a_real_excel_table(tmp_path):
    """Proves the Table object is really registered in the .xlsx, with
    banding — not just present in the in-memory openpyxl object.
    """
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    style = StyleManager.for_theme("premium")
    add_data_table(ws, style, ["Date", "Amount"], n_rows=20, table_name="Orders")
    out_path = wb.save(tmp_path / "orders.xlsx")

    raw = load_workbook(str(out_path))
    reloaded_ws = raw["Orders"]
    assert "Orders" in reloaded_ws.tables
    table = reloaded_ws.tables["Orders"]
    assert table.ref == "A1:B21"
    assert table.tableStyleInfo.showRowStripes is True
    assert reloaded_ws.freeze_panes == "A2"
