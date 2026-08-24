import pytest

from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.data.tables import ColumnSchema, add_typed_table
from excel_engine.navigation import internal_link
from excel_engine.protection import apply_standard_protection
from excel_engine.styles.style_manager import StyleManager
from excel_engine.validation.integrity_validator import validate_integrity, validate_protection


def test_no_duplicate_table_names_passes():
    wb = ExcelWorkbook()
    style = StyleManager.for_theme("premium")
    ws1 = wb.add_sheet("Orders")
    ws2 = wb.add_sheet("Income")
    add_typed_table(ws1, style, [ColumnSchema(header="Amount")], n_rows=5, table_name="OrdersTbl")
    add_typed_table(ws2, style, [ColumnSchema(header="Amount")], n_rows=5, table_name="IncomeTbl")
    assert validate_integrity(wb).passed is True


def test_duplicate_table_names_across_sheets_fails():
    """openpyxl's own `add_table` already refuses a duplicate name when
    building fresh via this engine — so simulate the one place this can
    still happen: a workbook loaded from an externally-edited/malformed
    .xlsx that already contains the duplicate.
    """
    from openpyxl.worksheet.table import Table

    wb = ExcelWorkbook()
    ws1 = wb.add_sheet("Orders")
    ws2 = wb.add_sheet("Income")
    ws1.raw["A1"] = "Amount"
    ws2.raw["A1"] = "Amount"
    ws1.raw.tables["SameName"] = Table(displayName="SameName", ref="A1:A1")
    ws2.raw.tables["SameName"] = Table(displayName="SameName", ref="A1:A1")

    result = validate_integrity(wb)
    assert result.passed is False
    assert "SameName" in result.issues[0].message


def test_add_table_itself_refuses_duplicate_names():
    """Documents the complementary guarantee: this engine's own build
    path can't produce the scenario above in the first place.
    """
    wb = ExcelWorkbook()
    style = StyleManager.for_theme("premium")
    ws1 = wb.add_sheet("Orders")
    ws2 = wb.add_sheet("Income")
    add_typed_table(ws1, style, [ColumnSchema(header="Amount")], n_rows=5, table_name="SameName")
    with pytest.raises(ValueError, match="already exists"):
        add_typed_table(ws2, style, [ColumnSchema(header="Amount")], n_rows=5, table_name="SameName")


def test_valid_hyperlinks_pass():
    wb = ExcelWorkbook()
    dashboard = wb.add_sheet("Dashboard")
    wb.add_sheet("Orders")
    dashboard.raw["A1"].hyperlink = internal_link(wb, "Orders")
    assert validate_integrity(wb).passed is True


def test_dangling_hyperlink_fails():
    wb = ExcelWorkbook()
    dashboard = wb.add_sheet("Dashboard")
    dashboard.raw["A1"].hyperlink = "#Goals!A1"
    result = validate_integrity(wb)
    assert result.passed is False
    assert "Goals" in result.issues[0].message


def test_empty_workbook_passes_integrity():
    wb = ExcelWorkbook()
    wb.add_sheet("Dashboard")
    assert validate_integrity(wb).passed is True


def test_protection_locked_formula_cell_passes():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Dashboard")
    ws.set_formula("A1", "=1+1")  # stays locked, the default
    result = validate_protection(wb, expected_locked_formula_cells=[("Dashboard", "A1")])
    assert result.passed is True


def test_protection_unlocked_formula_cell_fails():
    from excel_engine.protection.cells import unlock_range

    wb = ExcelWorkbook()
    ws = wb.add_sheet("Dashboard")
    ws.set_formula("A1", "=1+1")
    unlock_range(ws, "A1")  # accidentally left editable
    result = validate_protection(wb, expected_locked_formula_cells=[("Dashboard", "A1")])
    assert result.passed is False


def test_protection_unlocked_input_range_passes():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    apply_standard_protection(ws, editable_ranges=["A2:C10"])
    result = validate_protection(wb, expected_unlocked_input_ranges=[("Orders", "A2:C10")])
    assert result.passed is True


def test_protection_locked_input_range_fails():
    wb = ExcelWorkbook()
    wb.add_sheet("Orders")
    # never unlocked -> stays locked, which is wrong for a declared input range
    result = validate_protection(wb, expected_unlocked_input_ranges=[("Orders", "A2:C10")])
    assert result.passed is False
