import pytest
from openpyxl import load_workbook

from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.data.named_ranges import (
    add_named_range,
    get_named_range_formula,
    list_named_ranges,
)
from excel_engine.exceptions.errors import WorkbookBuildError


def test_add_named_range():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Orders")
    ws.set_value("A2", 100)
    add_named_range(wb, "TotalAmount", sheet="Orders", cell_range="A2")
    assert "TotalAmount" in list_named_ranges(wb)


def test_get_named_range_formula_is_absolute():
    wb = ExcelWorkbook()
    wb.add_sheet("Orders")
    add_named_range(wb, "TotalAmount", sheet="Orders", cell_range="A2")
    formula = get_named_range_formula(wb, "TotalAmount")
    assert formula == "Orders!$A$2"


def test_invalid_name_rejected():
    wb = ExcelWorkbook()
    wb.add_sheet("Orders")
    with pytest.raises(WorkbookBuildError):
        add_named_range(wb, "Not A Valid Name!", sheet="Orders", cell_range="A2")


def test_missing_named_range_raises():
    wb = ExcelWorkbook()
    wb.add_sheet("Orders")
    with pytest.raises(WorkbookBuildError):
        get_named_range_formula(wb, "DoesNotExist")


def test_named_range_survives_save_and_reload_and_is_usable_in_a_formula(tmp_path):
    wb = ExcelWorkbook()
    orders = wb.add_sheet("Orders")
    orders.set_value("A2", 100)
    add_named_range(wb, "TotalAmount", sheet="Orders", cell_range="A2")

    dashboard = wb.add_sheet("Dashboard")
    dashboard.set_formula("A1", "=SUM(TotalAmount)")

    out_path = wb.save(tmp_path / "test.xlsx")
    raw = load_workbook(str(out_path))
    assert "TotalAmount" in raw.defined_names
    assert raw["Dashboard"]["A1"].value == "=SUM(TotalAmount)"
