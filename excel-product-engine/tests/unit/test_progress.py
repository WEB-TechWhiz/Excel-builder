from openpyxl import load_workbook

from excel_engine.components.progress import add_progress_bar
from excel_engine.core.workbook import ExcelWorkbook
from excel_engine.styles.style_manager import StyleManager


def test_progress_bar_registers_conditional_formatting():
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Goals")
    style = StyleManager.for_theme("premium")
    add_progress_bar(ws, style, "B2:B10")

    ranges = [str(r.sqref) for r in ws.raw.conditional_formatting]
    assert "B2:B10" in ranges


def test_progress_bar_survives_save_and_reload(tmp_path):
    wb = ExcelWorkbook()
    ws = wb.add_sheet("Goals")
    style = StyleManager.for_theme("premium")
    ws.set_value("B2", 0.6)
    add_progress_bar(ws, style, "B2:B2")
    out_path = wb.save(tmp_path / "goals.xlsx")

    raw = load_workbook(str(out_path))
    ranges = [str(r.sqref) for r in raw["Goals"].conditional_formatting]
    # a true single-cell range round-trips through the .xlsx as "B2", not
    # "B2:B2" — Excel normalizes away the redundant colon on save/reload.
    assert "B2" in ranges
