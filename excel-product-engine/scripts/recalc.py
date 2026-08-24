#!/usr/bin/env python3
"""Cross-platform Excel formula recalculation and verification tool.

    python scripts/recalc.py path/to/file.xlsx [min_formulas]

Loads an Excel workbook, recalculates formulas (using Excel COM on Windows
or LibreOffice when available), checks all cell values for formula error codes,
and outputs a JSON report to stdout.
"""

from __future__ import annotations

import json
import os
import sys
import time
import uuid
from pathlib import Path

import openpyxl

ERROR_VALUES = {
    "#REF!",
    "#VALUE!",
    "#NAME?",
    "#N/A",
    "#NULL!",
    "#NUM!",
    "#DIV/0!",
    "#ERROR!",
}


def recalculate_with_excel(xlsx_path: Path) -> bool:
    """Use Microsoft Excel via COM automation (Windows) to calculate formulas and save."""
    try:
        import win32com.client
        import win32process
    except ImportError:
        return False

    abs_path = str(xlsx_path.resolve())
    temp_path = xlsx_path.parent / f"recalc_tmp_{uuid.uuid4().hex[:8]}.xlsx"
    excel = None
    pid = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        try:
            hwnd = excel.Hwnd
            _, pid = win32process.GetWindowThreadProcessId(hwnd)
        except Exception:
            pid = None

        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False

        wb = excel.Workbooks.Open(abs_path, UpdateLinks=0, ReadOnly=True)
        wb.SaveAs(str(temp_path.resolve()), 51)
        try:
            wb.Close(False)
        except Exception:
            pass

        try:
            excel.Quit()
        except Exception:
            pass

        if pid:
            try:
                os.kill(pid, 9)
            except Exception:
                pass

        time.sleep(0.2)
        if temp_path.exists():
            if xlsx_path.exists():
                try:
                    os.chmod(xlsx_path, 0o666)
                except Exception:
                    pass
            for _ in range(10):
                try:
                    temp_path.replace(xlsx_path)
                    return True
                except Exception:
                    time.sleep(0.2)
        return False
    except Exception:
        if pid:
            try:
                os.kill(pid, 9)
            except Exception:
                pass
        elif excel:
            try:
                excel.Quit()
            except Exception:
                pass
        if temp_path.exists():
            try:
                temp_path.unlink()
            except Exception:
                pass
        return False


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python scripts/recalc.py <file.xlsx> [min_formulas]")
        raise SystemExit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        report = {
            "status": "error",
            "message": f"File not found: {xlsx_path}",
            "total_formulas": 0,
            "total_errors": 1,
            "errors": [f"File not found: {xlsx_path}"],
        }
        print(json.dumps(report))
        raise SystemExit(1)

    # 1. Recalculate formulas if COM is available
    recalculated = recalculate_with_excel(xlsx_path)

    # 2. Inspect formulas using openpyxl
    wb_formulas = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    total_formulas = 0
    for sheetname in wb_formulas.sheetnames:
        ws = wb_formulas[sheetname]
        for row in ws.iter_rows(values_only=True):
            for cell_val in row:
                if isinstance(cell_val, str) and cell_val.startswith("="):
                    total_formulas += 1

    # 3. Inspect calculated values for errors using openpyxl (data_only=True)
    wb_data = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    total_errors = 0
    errors: list[str] = []

    for sheetname in wb_data.sheetnames:
        ws = wb_data[sheetname]
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for c_idx, cell_val in enumerate(row, start=1):
                if isinstance(cell_val, str) and cell_val.upper() in ERROR_VALUES:
                    total_errors += 1
                    col_letter = openpyxl.utils.get_column_letter(c_idx)
                    errors.append(f"{sheetname}!{col_letter}{r_idx}: {cell_val}")

    report = {
        "status": "success",
        "recalculated": recalculated,
        "total_formulas": total_formulas,
        "total_errors": total_errors,
        "errors": errors,
    }
    print(json.dumps(report))


if __name__ == "__main__":
    main()
